from app.models.bracket import *
from app.models.network import Network
from app.models.networkbranch import NetworkBranch
from app.models.networknode import NetworkNode
from app.models.networkuser import NetworkUser


class Serialization:

    @staticmethod
    def load_network_from_XLSX(file):
        """
        This method will parse a file .xls_ to create a network.
        :param file: the path of the file containing the network. Must be of ".xlsx" extension.
        :return: network: the one loaded from the Excel file.
        """
        from openpyxl import load_workbook
        from openpyxl.cell import cell
        network = Network()
        # Select a workbook
        workbook = load_workbook(file, data_only=True)
        worksheet = workbook.worksheets[0]
        # Loading from file: we will iterate over each row/column, considering the XL file as a matrix
        # And then creating nodes to add them to the network.
        # Row Counter
        i = 0
        # Values of all the rows
        full_nodes = []
        row_count = worksheet.max_row
        col_count = worksheet.max_column
        max_col = cell.get_column_letter(col_count)
        range = "A1:"+max_col+str(row_count)
        # Iterate over XL file
        for rowOfCellObjects in worksheet[range]:
            i += 1
            # Column Counter
            j = 0
            # Values of the current row
            row_node = []
            for cellObj in rowOfCellObjects:
                j += 1
                # If we are on the first row - Which means we will be reading column headers, we pass to the next row
                if i == 1 and j == 1:
                    break
                row_node.append(cellObj.value)
            if i != 1 and j != 1:
                # if not, we add the row to the stack that contains our XL rows
                full_nodes.append(row_node)
        # Now that we've had retrieved the data into an array, we will iterate over it to create a full network
        # First, we instanciate a new Bracket
        for node_in_file in full_nodes:
            node = NetworkNode()
            branch = NetworkBranch()
            user = NetworkUser([1,1,1], None, [0,0,0])
            # We then set its bracket's ID.
            node.set_id(int(node_in_file[0][5:]))
            # Same as the parent.
            node.set_parent_id(int(node_in_file[1]))
            # Now, we're on to create the Branch.
            # First, we're setting the phases that's on the branch.
            branch.set_phases(node_in_file[2:5])
            # Then, the representation types (For later, that's why it's in comment)
            # bracket.set_types(node_in_file[5:8])
            # Setting the length
            branch.set_length(node_in_file[8])
            # Finally, we instanciate R and X: vectors of floating point numbers.
            # Their structure is as follows
            # ## Note that we have 4 values for each possible phases: A, B, C and N.
            # R = [RA, RB, RC, RD]
            # Examples: RA = [R_AA, R_AB, R_AC, R_AN]
            import math
            RA = [node_in_file[9], node_in_file[15], node_in_file[17], math.inf]
            RB = [node_in_file[15], node_in_file[11], node_in_file[19], math.inf]
            RC = [node_in_file[17], node_in_file[19], node_in_file[13], math.inf]
            RN = [math.inf, math.inf, math.inf, math.inf]
            branch.set_R([RA, RB, RC, RN])
            # Same as for R
            XA = [node_in_file[10], node_in_file[16], node_in_file[18], math.inf]
            XB = [node_in_file[16], node_in_file[12], node_in_file[20], math.inf]
            XC = [node_in_file[18], node_in_file[20], node_in_file[14], math.inf]
            XN = [math.inf, math.inf, math.inf, math.inf]
            branch.set_X([XA, XB, XC, XN])
            """bracket.set_index(node_in_file[21:24])
            bracket.set_EAN(node_in_file[24:27])
            bracket.set_EAN(node_in_file[27:29])"""
            # Now we are setting the Power that circulates through the branch.
            pTest = node_in_file[30]
            powers = []
            powers.append(pTest)
            user.set_P(powers)
            node.set_users([user, user, user])
            if node_in_file[1] != 0:
                node.set_parent(network.find_node(node_in_file[1]))
            bracket = Bracket(node, branch)
            network.add_bracket(bracket)
            # if we are not on the first real node of the network
            if bracket.get_node().get_id() != 1:
                network.set_parent_of_bracket(bracket, node.get_parent_id()-1)
        # Return the network created from the XL file.
        return network


if __name__ == '__main__':
    pass